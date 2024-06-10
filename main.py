from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('700x412+300+200')
root.resizable(False,False)
root.configure(bg="#5277ff")
file=pathlib.Path('data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1'] = "Full name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save('data.xlsx')

def clear():
    NameValue.set('')
    ContactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)

def submit():
    name = NameValue.get()
    contact = ContactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0,END)

    file = openpyxl.load_workbook('data.xlsx')
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1,
               value=name)
    sheet.cell(column=2, row=sheet.max_row,
               value=contact)
    sheet.cell(column=3, row=sheet.max_row ,
               value=age)
    sheet.cell(column=4, row=sheet.max_row ,
               value=gender)
    sheet.cell(column=5, row=sheet.max_row ,
               value=address)

    file.save(r'data.xlsx')
    messagebox.showinfo('info','details added!')
    NameValue.set('')
    ContactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)

#heading
Label(root, text="Please fill out this form:",
      font="calibri 14 bold", bg="#5277ff", fg="#fff")\
    .place(x=50, y=20)

#label
Label(root, text='Name', font="calibri 14 bold", bg="#5277ff",
      fg="#fff").place(x=50, y=100)
Label(root, text='Phone Number', font="calibri 14 bold", bg="#5277ff",
      fg="#fff").place(x=50, y=150)
Label(root, text='Age', font="calibri 14 bold", bg="#5277ff",
      fg="#fff").place(x=50, y=200)
Label(root, text='Gender', font="calibri 14 bold", bg="#5277ff",
      fg="#fff").place(x=360, y=200)
Label(root, text='Address', font="calibri 14 bold", bg="#5277ff",
      fg="#fff").place(x=50, y=250)

#entry
NameValue= StringVar()
ContactValue= StringVar()
AgeValue= StringVar()

nameEntry=Entry(root,textvariable=NameValue,width=45,
                bd=2,font=20)
contactEntry=Entry(root,
                textvariable=ContactValue,width=45,
                bd=2,font=20)
ageEntry=Entry(root,textvariable=AgeValue,width=15,
                bd=2,font=20)
nameEntry.place(x=200, y=100)
contactEntry.place(x=200, y=150)
ageEntry.place(x=200, y=200)

#gender
gender_combobox=Combobox(root, values=['Male','Female'],
                    font='calibri 14 bold',state='r',width=14)
gender_combobox.place(x=440, y=200)
gender_combobox.set('Male')

#address
addressEntry=Text(root,width=50,height=4,bd=4)
addressEntry.place(x=200,y=250)

#Buttons
Button(root,text='Submit',bg='#5e17eb',fg='white',
       font='calibri 10 bold', width=15 ,height=2,
       command= submit).place(x=200,y=350)

Button(root,text='Clear',bg='#5e17eb',fg='white',
       font='calibri 10 bold',width=15, height=2,
       command= clear).place(x=340,y=350)

Button(root,text='Exit',bg='#5e17eb',fg='white',
       font='calibri 10 bold',width=15, height=2,
       command= lambda:root.destroy()).place(x=480,y=350)

root.mainloop()
